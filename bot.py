# BU YERDA TO‘LIQ bot.py KODI TURADI
# Sizga oldin tayyorlab bergan funksional: test yuklash, premium/oddiy ajratish, GitHub yuklash, foydalanuvchi boshqaruvi va h.k.
# Kodingizni ilova qilib qo‘yaman
import logging
import os
import openpyxl
import sqlite3
import asyncio
import datetime
import requests

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
    CallbackQueryHandler,
)

# --- Loglarni sozlash ---
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

# --- Holatlarni aniqlash ---
CHOOSING_TEST, ANSWERING = range(2)

# --- Fayl va papkalar ---
TESTS_FOLDER = "tests"
DB_NAME = "bot_data.db"
GITHUB_TOKEN = "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"
GITHUB_REPO = "Zarifjon-dev/tarixiy-test-bot"
GITHUB_BRANCH = "main"

ADMIN_USERNAME = "@Tarix_dtmadmin"

# --- Global testlar ---
tests_data = {"oddiy": [], "premium": []}

# --- Ma'lumotlar bazasi ---
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            current_question_index INTEGER,
            score INTEGER,
            is_premium INTEGER DEFAULT 0
        )
    ''')
    conn.commit()
    conn.close()

def get_user_data(user_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT current_question_index, score, is_premium FROM users WHERE user_id = ?", (user_id,))
    result = cursor.fetchone()
    conn.close()
    if result:
        return {"current_question_index": result[0], "score": result[1], "is_premium": result[2]}
    return None

def save_user_data(user_id, current_question_index, score, is_premium=0):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO users (user_id, current_question_index, score, is_premium)
        VALUES (?, ?, ?, ?)
    ''', (user_id, current_question_index, score, is_premium))
    conn.commit()
    conn.close()

def set_premium(user_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET is_premium = 1 WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

def delete_user(user_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

# --- Testlarni yuklash ---
def load_tests():
    for filename in os.listdir(TESTS_FOLDER):
        if filename.endswith(".xlsx"):
            full_path = os.path.join(TESTS_FOLDER, filename)
            test_type = "premium" if "premium" in filename.lower() else "oddiy"
            workbook = openpyxl.load_workbook(full_path)
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                question = sheet.cell(row=row, column=1).value
                options = [sheet.cell(row=row, column=i).value for i in range(2, 6)]
                correct = sheet.cell(row=row, column=6).value
                if question and correct:
                    tests_data[test_type].append({
                        "savol": question,
                        "variantlar": [opt for opt in options if opt],
                        "togri_javob": correct
                    })
    logger.info("Testlar yuklandi")

# --- GitHub yuklash ---
def upload_to_github(file_path, filename):
    with open(file_path, "rb") as f:
        content = f.read()
    import base64
    encoded = base64.b64encode(content).decode("utf-8")
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/tests/{filename}"
    headers = {"Authorization": f"token {GITHUB_TOKEN}"}
    data = {
        "message": f"Uploaded {filename}",
        "content": encoded,
        "branch": GITHUB_BRANCH
    }
    r = requests.put(url, headers=headers, json=data)
    logger.info(f"GitHubga yuklandi: {r.status_code}")

# --- Bot funksiyalari ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Test turini tanlang:",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Oddiy testlar", callback_data="oddiy")],
            [InlineKeyboardButton("Premium testlar", callback_data="premium")],
            [InlineKeyboardButton("Pro versiyaga o‘tish", url=f"https://t.me/{Tarix_dtmadmin.lstrip('@')}")]
        ])
    )
    return CHOOSING_TEST

async def choose_test(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    mode = query.data

    is_premium = get_user_data(user_id) or {"is_premium": 0}
    if mode == "premium" and not is_premium["is_premium"]:
        await query.edit_message_text("Pro versiyaga o‘tish uchun adminga yozing: " + ADMIN_USERNAME)
        return ConversationHandler.END

    delete_user(user_id)
    save_user_data(user_id, 0, 0, 1 if mode == "premium" else 0)
    await query.edit_message_text("Test boshlandi")
    return await ask_question(update, context, mode)

async def ask_question(update: Update, context: ContextTypes.DEFAULT_TYPE, mode=None):
    user_id = update.effective_user.id
    data = get_user_data(user_id)
    if not data:
        return ConversationHandler.END
    index = data["current_question_index"]
    test_type = "premium" if data["is_premium"] else "oddiy"

    if index >= len(tests_data[test_type]):
        await update.effective_message.reply_text(
            f"Test tugadi. Natija: {data['score']} / {len(tests_data[test_type])}"
        )
        delete_user(user_id)
        return ConversationHandler.END

    q = tests_data[test_type][index]
    keyboard = [[InlineKeyboardButton(opt, callback_data=opt)] for opt in q["variantlar"]]
    await update.effective_message.reply_text(
        f"{index+1}. {q['savol']}",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ANSWERING

async def handle_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    data = get_user_data(user_id)
    index = data["current_question_index"]
    test_type = "premium" if data["is_premium"] else "oddiy"
    question = tests_data[test_type][index]

    if query.data == question["togri_javob"]:
        data["score"] += 1
        feedback = "✅ To‘g‘ri!"
    else:
        feedback = f"❌ Noto‘g‘ri. To‘g‘ri javob: {question['togri_javob']}"

    await query.edit_message_text(query.message.text + "\n\n" + feedback)
    data["current_question_index"] += 1
    save_user_data(user_id, data["current_question_index"], data["score"], data["is_premium"])
    await asyncio.sleep(0.5)
    return await ask_question(update, context)

async def upload_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if user.username != ADMIN_USERNAME.lstrip("@"): return

    if not update.message.document:
        await update.message.reply_text("Excel fayl yuboring")
        return

    doc = update.message.document
    file = await doc.get_file()
    filename = doc.file_name
    path = os.path.join(TESTS_FOLDER, filename)
    await file.download_to_drive(path)
    upload_to_github(path, filename)
    await update.message.reply_text(f"{filename} yuklandi va GitHubga qo‘shildi")
    load_tests()

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    delete_user(update.effective_user.id)
    await update.message.reply_text("Test bekor qilindi")
    return ConversationHandler.END

def main():
    init_db()
    load_tests()
    TOKEN = os.environ.get("7775497614:AAFRrodSyDotYX0AMIG7o0ijMXXizcSsbxg")
    application = Application.builder().token(7775497614:AAFRrodSyDotYX0AMIG7o0ijMXXizcSsbxg).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            CHOOSING_TEST: [CallbackQueryHandler(choose_test)],
            ANSWERING: [CallbackQueryHandler(handle_answer)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )

    application.add_handler(conv_handler)
    application.add_handler(MessageHandler(filters.Document.ALL, upload_file))
    application.run_polling()

if __name__ == "__main__":
    main()
